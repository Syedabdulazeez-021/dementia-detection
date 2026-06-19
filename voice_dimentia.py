# -*- coding: utf-8 -*-
"""
Dementia Analysis System  v4  — Fast Edition
=============================================
Speed fixes vs previous version:
  • librosa.pyin (slow Viterbi)  →  librosa.yin   (~8x faster)
  • Python filter-bank loops     →  fully vectorised numpy  (~20x faster)
  • matplotlib backend set to Agg (no GUI dependency, instant save)
  • Removed redundant double-extraction in ensemble path
  • Pre-emphasis applied with np.diff (no Python loop)

v4.1 additional speed fixes (post-recording):
  • HNR: np.correlate O(n²) on 1.2M samples → FFT autocorrelation O(n log n) on 2s clip (~100x)
  • Jitter: librosa.yin on full 55s → first 15s only (~3.5x)
  • Shimmer: vectorised on full 55s → first 10s only (~5.5x)
  • n_mfcc 42 → 30  (only indices 0,1,15,29 are used; ~28% faster MFCC)
  • Feature clipping: Python for-loop → single vectorised np.clip call

Retained:
  • Live recording progress bar + mic level meter
  • Waveform + envelope + F0 plot  (waveform_plot.png)
  • Feature bar chart + radar + prediction summary  (feature_plot.png)
  • Excel export of top-8 discriminant features only
  • Indian-accent calibration + 60% confidence threshold
  • 2-segment ensemble for long audio

Usage:
  python voice_dimentia.py --audio input.wav --plot --show-features
  python voice_dimentia.py  (microphone)
"""

import sys, io, os, warnings
from datetime import datetime

warnings.filterwarnings('ignore')
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import numpy as np
import librosa
import pickle
from scipy.io.wavfile import write as wav_write
from scipy.stats import kurtosis, skew

try:
    import sounddevice as sd
    SD_AVAILABLE = True
except ImportError:
    SD_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import matplotlib
matplotlib.use('Agg')          # headless — no Tk/Qt needed, fastest option
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec

try:
    from surfboard.sound import Waveform
    SURFBOARD_AVAILABLE = True
except ImportError:
    SURFBOARD_AVAILABLE = False

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
TRAINING_STATS = {
    'MFCC2':          (197.9443, 29.9131,   62.8883,  268.9908),
    'kurt_MFCC30':    (  2.0787,  2.6332,   -1.3718,   28.0893),
    'mean_MFCC30':    ( -0.5763,  2.0878,   -7.7479,   14.0492),
    'skew_MFCC2':     (  0.2464,  0.8634,   -4.9125,    3.9258),
    'mean_MFCC16':    ( -2.3350,  4.0140,  -20.3263,   11.4448),
    'flt_bnk_eng22':  ( 20.4155, 627.5743,-1944.8118, 2140.0957),
    'MFCC30':         ( 24.5920,  8.2870,   -1.5456,   60.4300),
    'kurt_MFCC16':    (  1.0946,  1.7921,   -1.2614,   15.6081),
    'flt_bnk_eng2':   ( 27.1581, 627.1720,-1735.9043, 2639.4044),
    'flt_bnk_eng24':  ( 19.2840, 627.0900,-1921.5678, 2157.2382),
    'MFCC1':          (-245.4195, 25.8512, -361.9381,   77.6719),
    'flt_bnk_eng15':  ( 18.9957, 627.2461,-1952.2539, 2213.4082),
    'kurt_MFCC2':     (  0.4722,  2.6890,   -1.6476,   30.4984),
    'flt_bnk_eng20':  ( 20.5133, 627.3841,-1928.6832, 2098.4675),
    'flt_bnk_eng13':  ( 18.9423, 627.4219,-1879.3162, 2294.0684),
    'n_sil_segments': (  3.1715,  1.5466,    1.0000,   10.0000),
    'frac_silence':   ( 34.5341, 16.7483,    0.0023,   85.2797),
    'min_sil_len':    (  1.4550,  1.0143,    0.0000,   12.3348),
    'jitter':         (  0.0083,  0.0036,    0.0000,    0.0424),
    'shimmer':        (  0.6766,  0.1632,    0.0000,    3.6390),
    'HNR':            (  8.8838,  2.9212,    0.0000,   22.6416),
}

AD_MEANS = {
    'MFCC2': 198.589, 'kurt_MFCC30': 2.558, 'mean_MFCC30': -0.376,
    'skew_MFCC2': 0.439, 'mean_MFCC16': -2.137, 'flt_bnk_eng22': 31.324,
    'MFCC30': 25.326, 'kurt_MFCC16': 1.587, 'flt_bnk_eng2': 20.321,
    'flt_bnk_eng24': 31.203, 'MFCC1': -236.775, 'flt_bnk_eng15': 19.801,
    'kurt_MFCC2': 0.976, 'flt_bnk_eng20': 28.420, 'flt_bnk_eng13': 19.502,
    'n_sil_segments': 3.354, 'frac_silence': 37.426, 'min_sil_len': 1.475,
    'jitter': 0.009, 'shimmer': 0.691, 'HNR': 9.246,
}
CN_MEANS = {
    'MFCC2': 197.355, 'kurt_MFCC30': 1.641, 'mean_MFCC30': -0.759,
    'skew_MFCC2': 0.070, 'mean_MFCC16': -2.516, 'flt_bnk_eng22': 10.443,
    'MFCC30': 23.921, 'kurt_MFCC16': 0.645, 'flt_bnk_eng2': 33.409,
    'flt_bnk_eng24': 8.388, 'MFCC1': -253.322, 'flt_bnk_eng15': 18.260,
    'kurt_MFCC2': 0.011, 'flt_bnk_eng20': 13.285, 'flt_bnk_eng13': 18.431,
    'n_sil_segments': 3.005, 'frac_silence': 31.890, 'min_sil_len': 1.437,
    'jitter': 0.008, 'shimmer': 0.664, 'HNR': 8.552,
}

FEATURE_ORDER = [
    'MFCC2', 'kurt_MFCC30', 'mean_MFCC30', 'skew_MFCC2', 'mean_MFCC16',
    'flt_bnk_eng22', 'MFCC30', 'kurt_MFCC16', 'flt_bnk_eng2', 'flt_bnk_eng24',
    'MFCC1', 'flt_bnk_eng15', 'kurt_MFCC2', 'flt_bnk_eng20', 'flt_bnk_eng13',
    'n_sil_segments', 'frac_silence', 'min_sil_len', 'jitter', 'shimmer', 'HNR',
]

# Top-8 by Cohen's d from training data
EXCEL_FEATURES = [
    'kurt_MFCC16',   # d=0.574
    'skew_MFCC2',    # d=0.438
    'kurt_MFCC2',    # d=0.379
    'kurt_MFCC30',   # d=0.356
    'jitter',        # d=0.352
    'frac_silence',  # d=0.335
    'HNR',           # d=0.240
    'shimmer',       # d=0.166
]

FEATURE_DESCRIPTIONS = {
    'kurt_MFCC16':  'Spectral Kurtosis MFCC16  d=0.574',
    'skew_MFCC2':   'Spectral Skewness MFCC2   d=0.438',
    'kurt_MFCC2':   'Spectral Kurtosis MFCC2   d=0.379',
    'kurt_MFCC30':  'Spectral Kurtosis MFCC30  d=0.356',
    'jitter':       'Voice Jitter              d=0.352',
    'frac_silence': 'Silence Fraction (%)      d=0.335',
    'HNR':          'Harmonics-to-Noise Ratio  d=0.240',
    'shimmer':      'Voice Shimmer             d=0.166',
}

INDIAN_ACCENT_OFFSETS = {
    'MFCC2':          -0.20, 'kurt_MFCC30':    -0.10, 'mean_MFCC30':    -0.12,
    'skew_MFCC2':     -0.15, 'mean_MFCC16':    -0.14, 'flt_bnk_eng22':  -0.10,
    'MFCC30':         -0.12, 'kurt_MFCC16':    -0.20, 'flt_bnk_eng2':   -0.06,
    'flt_bnk_eng24':  -0.10, 'MFCC1':           0.12, 'flt_bnk_eng15':  -0.08,
    'kurt_MFCC2':     -0.18, 'flt_bnk_eng20':  -0.10, 'flt_bnk_eng13':  -0.08,
    'n_sil_segments':  0.20, 'frac_silence':   -0.08, 'min_sil_len':     0.25,
    'jitter':         -0.22, 'shimmer':        -0.10, 'HNR':             0.18,
}

# Friendly names for voice features (used by the explainability panel/report)
VOICE_FEATURE_LABELS = {
    'jitter':        'Voice tremor (jitter)',
    'shimmer':       'Amplitude instability (shimmer)',
    'HNR':           'Harmonics-to-noise ratio',
    'frac_silence':  'Silence fraction',
    'n_sil_segments':'Pause count',
    'min_sil_len':   'Min pause length',
    'skew_MFCC2':    'Spectral skew (MFCC2)',
    'kurt_MFCC2':    'Spectral kurtosis (MFCC2)',
    'kurt_MFCC16':   'Spectral kurtosis (MFCC16)',
    'kurt_MFCC30':   'Spectral kurtosis (MFCC30)',
    'mean_MFCC16':   'Spectral mean (MFCC16)',
    'mean_MFCC30':   'Spectral mean (MFCC30)',
    'MFCC1':         'Spectral coeff. (MFCC1)',
    'MFCC2':         'Spectral coeff. (MFCC2)',
    'MFCC30':        'Spectral coeff. (MFCC30)',
}


def log_session_for_retraining(cal_vector, prediction, proba, source='gui',
                               path='retrain_data.csv'):
    """
    Append this session's full 21-feature vector + the model's prediction to a
    CSV, leaving a blank `true_label` column. When a clinician later CONFIRMS
    the real diagnosis, they fill in true_label (0 = healthy, 1 = dementia),
    and retrain.py can then learn from those confirmed cases.

    This is the honest 'learn from feedback' mechanism: the model improves from
    confirmed ground-truth labels, never from its own unverified guesses.
    Fully guarded so it can never disrupt analysis.
    """
    try:
        import csv as _csv
        vec = list(cal_vector)
        if len(vec) != len(FEATURE_ORDER):
            return
        exists = os.path.exists(path)
        with open(path, 'a', newline='') as f:
            w = _csv.writer(f)
            if not exists:
                w.writerow(['timestamp', 'source', 'predicted_label', 'p_ad']
                           + list(FEATURE_ORDER) + ['true_label'])
            w.writerow([datetime.now().strftime('%Y-%m-%d %H:%M:%S'), source,
                        int(prediction), round(float(proba[1]), 4)]
                       + [round(float(v), 6) for v in vec] + [''])
    except Exception:
        pass


def _compute_voice_attribution(model, cal_vector, top_n=5):
    """
    Approximate per-sample feature attribution for the voice Random Forest.

    Combines the model's global feature_importances_ with how 'AD-like' each
    feature is for THIS recording (its position between the training CN mean
    and AD mean). Returns the top-N drivers as fractions that sum to 1.0, so
    the explainability panel can split the voice contribution across them.

    Fully defensive: returns [] on any problem, so the app falls back to a
    single combined 'Voice risk' contribution.
    """
    try:
        importances = getattr(model, 'feature_importances_', None)
        if importances is None or len(importances) != len(FEATURE_ORDER):
            return []
        vec = list(cal_vector)
        saliences = []
        for i, name in enumerate(FEATURE_ORDER):
            ad = AD_MEANS.get(name)
            cn = CN_MEANS.get(name)
            if ad is None or cn is None:
                continue
            denom = (ad - cn)
            if abs(denom) < 1e-9:
                continue
            # 0 = at CN (healthy) mean, 1 = at AD mean; clip to [0, 1]
            ad_likeness = (vec[i] - cn) / denom
            ad_likeness = max(0.0, min(1.0, ad_likeness))
            saliences.append((name, float(importances[i]) * ad_likeness))
        saliences = [s for s in saliences if s[1] > 0]
        if not saliences:
            return []
        saliences.sort(key=lambda x: x[1], reverse=True)
        top = saliences[:top_n]
        total = sum(s[1] for s in top)
        if total <= 0:
            return []
        return [{'name': n,
                 'label': VOICE_FEATURE_LABELS.get(n, n),
                 'fraction': round(s / total, 4)} for n, s in top]
    except Exception:
        return []


# =============================================================================
class DementiaAnalyser:

    def __init__(self, model_path='dementia_rf_model.pkl', scaler_path='scaler.pkl',
                 indian_accent=True, confidence_threshold=0.60, output_dir='.'):
        self.model_path           = model_path
        self.scaler_path          = scaler_path
        self.sample_rate          = 22050
        self.duration             = 55
        self.indian_accent        = indian_accent
        self.confidence_threshold = confidence_threshold
        self.output_dir           = output_dir
        self._feat_idx            = {fn: i for i, fn in enumerate(FEATURE_ORDER)}
        # Pre-build clip arrays once (used in extract_features)
        self._clip_lo = np.array([TRAINING_STATS[fn][2] for fn in FEATURE_ORDER])
        self._clip_hi = np.array([TRAINING_STATS[fn][3] for fn in FEATURE_ORDER])
        self._load_model()

    # ── model ─────────────────────────────────────────────────────────────────
    def _load_model(self):
        try:
            import joblib
            self.model  = joblib.load(self.model_path)
            self.scaler = joblib.load(self.scaler_path)
            print("[OK] Model + Scaler loaded (joblib)")
        except Exception:
            with open(self.model_path,  'rb') as f: self.model  = pickle.load(f)
            with open(self.scaler_path, 'rb') as f: self.scaler = pickle.load(f)
            print("[OK] Model + Scaler loaded (pickle)")

    # ── live recording with progress bar ─────────────────────────────────────
    def record_audio(self, duration=None):
        if not SD_AVAILABLE:
            raise ImportError("pip install sounddevice")
        import time

        duration  = duration or self.duration
        sr        = self.sample_rate
        n_samples = int(duration * sr)
        buf       = np.zeros(n_samples, dtype='float32')
        pos       = [0]

        def _cb(indata, frames, t, status):
            end = pos[0] + frames
            if end > n_samples: frames = n_samples - pos[0]; end = n_samples
            if frames > 0: buf[pos[0]:end] = indata[:frames, 0]
            pos[0] = min(end, n_samples)

        print("\n" + "=" * 60)
        print("  MICROPHONE RECORDING")
        print(f"  Duration : {duration} seconds")
        print("  Speak naturally — describe your day, a memory, a picture …")
        print("=" * 60)
        for i in (3, 2, 1):
            sys.stdout.write(f"\r  Starting in {i} …   ")
            sys.stdout.flush()
            time.sleep(1)
        sys.stdout.write("\r" + " " * 50 + "\r")
        print("\n  ██ RECORDING STARTED — PLEASE SPEAK NOW ██\n")
        sys.stdout.flush()

        stream = sd.InputStream(samplerate=sr, channels=1, dtype='float32',
                                callback=_cb, blocksize=512)
        stream.start()
        t0 = time.time()
        BW, MW = 30, 10

        try:
            while True:
                elapsed = time.time() - t0
                if elapsed >= duration: break
                filled   = int(BW * elapsed / duration)
                bar      = '█' * filled + '░' * (BW - filled)
                chunk    = max(1, int(0.5 * sr))
                snippet  = buf[max(0, pos[0] - chunk): pos[0]]
                rms_live = float(np.sqrt(np.mean(snippet ** 2))) if len(snippet) else 0.0
                mf       = min(MW, int(rms_live * MW / 0.08))
                mic_bar  = '█' * mf + '░' * (MW - mf)
                icon     = '🔴' if rms_live < 0.002 else '🟢'
                line     = (f"  [{bar}] {int(elapsed):>2}/{duration}s  "
                            f"MIC {icon} [{mic_bar}]  RMS={rms_live:.4f}   ")
                sys.stdout.write('\r' + line)
                sys.stdout.flush()
                time.sleep(0.25)
        finally:
            stream.stop(); stream.close()

        sys.stdout.write('\n')
        print("\n  ██ RECORDING COMPLETE ██\n")
        audio = buf[:pos[0]]
        rms = float(np.sqrt(np.mean(audio ** 2)))
        print(f"  Captured : {len(audio)/sr:.1f}s  |  RMS={rms:.5f}")
        if rms < 0.001: print("  [WARN] Very quiet — check microphone.")
        elif rms < 0.005: print("  [WARN] Low level — speak louder.")
        else: print("  [OK] Audio level good.")
        return audio

    # ── signal utilities ──────────────────────────────────────────────────────
    def _envelope(self, signal, interval=50):
        """Rolling max envelope — stride trick, zero copy."""
        n     = len(signal) - interval
        abs_s = np.abs(signal)
        shape   = (n, interval)
        strides = (abs_s.strides[0], abs_s.strides[0])
        mat = np.lib.stride_tricks.as_strided(abs_s, shape=shape, strides=strides)
        return mat.max(axis=1)

    def _silence_detection(self, env):
        n1, k1 = 0.95, 1.05
        th = k1 * np.mean(np.sort(env)[:int(len(env) * n1)])
        ind = np.argwhere(np.abs(env) > th).T[0]
        if len(ind) < 2:
            return np.array([0]), np.array([len(env)]), 1
        ind2 = np.roll(ind, -1); ind2[-1] = ind[-1]
        ind3 = np.roll(ind,  1); ind3[ 0] = ind[ 0]
        min_gap = int(0.70 * self.sample_rate)
        mask    = (np.abs(ind - ind2) > min_gap) | ((ind - ind3) > min_gap)
        ind_new = ind[mask]
        if len(ind_new) == 0:
            mx = np.argmax(np.abs(ind - ind2))
            ind_new = np.array([ind[mx], ind2[mx]])
        ss = ind_new[::2]; se = ind_new[1::2]
        return ss, se, min(len(ss), len(se))

    def _filter_banks(self, signal):
        """FAST vectorised Mel filter bank — ~20x faster than Python loops."""
        pre = 0.95
        emph = np.empty_like(signal)
        emph[0] = signal[0]
        emph[1:] = signal[1:] - pre * signal[:-1]

        fl   = int(round(0.025 * self.sample_rate))
        fst  = int(round(0.010 * self.sample_rate))
        sl   = len(emph)
        nf   = int(np.ceil((sl - fl) / fst))
        pad  = np.zeros(nf * fst + fl - sl)
        sig  = np.concatenate([emph, pad])

        shape   = (nf, fl)
        strides = (sig.strides[0] * fst, sig.strides[0])
        frames  = np.lib.stride_tricks.as_strided(sig, shape=shape, strides=strides).copy()
        frames *= np.hamming(fl)

        NFFT  = 512
        mag   = np.abs(np.fft.rfft(frames, NFFT))
        pw    = (mag ** 2) / NFFT

        nfilt   = 26
        hi_mel  = 2595 * np.log10(1 + self.sample_rate / 1400)
        mels    = np.linspace(0, hi_mel, nfilt + 2)
        hz      = 700 * (10 ** (mels / 2595) - 1)
        bins    = np.floor((NFFT + 1) * hz / self.sample_rate).astype(int)
        fbank   = np.zeros((nfilt, NFFT // 2 + 1))
        for m in range(1, nfilt + 1):
            lo, c, hi = bins[m-1], bins[m], bins[m+1]
            ku = np.arange(lo, c);   kd = np.arange(c, hi)
            if len(ku): fbank[m-1, ku] = (ku - lo) / (c - lo + 1e-8)
            if len(kd): fbank[m-1, kd] = (hi - kd) / (hi - c + 1e-8)

        fb = pw @ fbank.T
        fb = np.where(fb == 0, np.finfo(float).eps, fb)
        return 20 * np.log10(fb)

    # ── prosodic features ─────────────────────────────────────────────────────
    def _estimate_jitter(self, signal):
        """
        librosa.yin on first 15 s only — sufficient voiced frames for jitter;
        avoids running YIN on the full 55-second buffer (~3.5x speedup).
        """
        sig = signal[:int(15 * self.sample_rate)]   # [OPT] truncate to 15 s
        try:
            f0 = librosa.yin(sig,
                             fmin=librosa.note_to_hz('C2'),
                             fmax=librosa.note_to_hz('C7'),
                             sr=self.sample_rate,
                             frame_length=2048, hop_length=512)
            vf0 = f0[f0 > 65]
            if len(vf0) < 4: return 0.008
            periods = self.sample_rate / vf0
            return float(np.clip(np.std(periods) / (np.mean(periods) + 1e-8), 0, 0.05))
        except Exception:
            return 0.008

    def _estimate_shimmer(self, signal):
        """Amplitude jitter on first 10 s — ~5.5x speedup vs full buffer."""
        signal = signal[:int(10 * self.sample_rate)]   # [OPT] truncate to 10 s
        fl  = int(0.025 * self.sample_rate)
        hop = int(0.010 * self.sample_rate)
        n   = (len(signal) - fl) // hop
        idx = np.arange(n)[:, None] * hop + np.arange(fl)
        frames = signal[idx]
        rms = np.sqrt(np.mean(frames ** 2, axis=1) + 1e-10)
        if len(rms) < 4: return 0.5
        voiced = rms[rms > np.percentile(rms, 20)]
        if len(voiced) < 4: voiced = rms
        return float(np.clip(np.std(voiced) / (np.mean(voiced) + 1e-8), 0, 2.0))

    def _estimate_hnr(self, signal):
        """
        FFT-based autocorrelation on 2 s clip — O(n log n) instead of O(n²).
        Original np.correlate('full') on 1.2 M samples was the #1 bottleneck.
        Only the first 1000 lags are used, so 2 s is plenty (~100x speedup).
        """
        seg  = signal[:int(2 * self.sample_rate)]   # [OPT] 2 s clip
        s    = seg - seg.mean()
        n    = len(s)
        nfft = 2 ** int(np.ceil(np.log2(2 * n)))   # next power-of-2 → fastest FFT
        S    = np.fft.rfft(s, n=nfft)
        ac   = np.fft.irfft(S * np.conj(S))[:n].real   # [OPT] FFT autocorrelation
        ac  /= (ac[0] + 1e-8)
        if len(ac) > 1:
            pk = np.clip(ac[1:min(len(ac), 1000)].max(), 1e-6, 1 - 1e-6)
            return float(np.clip(10 * np.log10(pk / (1 - pk + 1e-8) + 1e-8), 0, 20))
        return 8.0

    def _prosodic(self, signal):
        if SURFBOARD_AVAILABLE:
            try:
                wav_write('_tmp.wav', self.sample_rate, signal)
                snd = Waveform(path='_tmp.wav', sample_rate=self.sample_rate)
                j = snd.jitters()['localJitter']
                s = snd.shimmers()['localdbShimmer']
                h = snd.hnr()
                os.remove('_tmp.wav')
                return j, s, h
            except Exception:
                pass
        return self._estimate_jitter(signal), self._estimate_shimmer(signal), self._estimate_hnr(signal)

    # ── feature extraction ────────────────────────────────────────────────────
    def extract_features(self, signal):
        print("[PROCESSING] Extracting features …", end=' ', flush=True)

        # [OPT] n_mfcc 42 → 30: only indices 0, 1, 15, 29 are used (~28% faster)
        MFCCs = librosa.feature.mfcc(y=signal, sr=self.sample_rate,
                                     n_fft=2048, hop_length=512, n_mfcc=30)
        fb    = self._filter_banks(signal)
        env   = self._envelope(signal)
        ss, se, _ = self._silence_detection(env)

        sil_sec = float(np.sum(se - ss)) / self.sample_rate
        sig_dur = len(env) / self.sample_rate
        sil_pct = (sil_sec / sig_dur * 100) if sig_dur > 0 else 0.0
        min_sil = float(np.min(se - ss)) / self.sample_rate if len(ss) else 0.0

        jitter, shimmer, hnr = self._prosodic(signal)

        raw = np.array([
            np.max(MFCCs[1, :]),          # MFCC2
            kurtosis(MFCCs[29, :]),        # kurt_MFCC30
            np.mean(MFCCs[29, :]),         # mean_MFCC30
            skew(MFCCs[1, :]),             # skew_MFCC2
            np.mean(MFCCs[15, :]),         # mean_MFCC16
            np.sum(fb[:, 21]),             # flt_bnk_eng22
            np.max(MFCCs[29, :]),          # MFCC30
            kurtosis(MFCCs[15, :]),        # kurt_MFCC16
            np.sum(fb[:, 1]),              # flt_bnk_eng2
            np.sum(fb[:, 23]),             # flt_bnk_eng24
            np.max(MFCCs[0, :]),           # MFCC1
            np.sum(fb[:, 14]),             # flt_bnk_eng15
            kurtosis(MFCCs[1, :]),         # kurt_MFCC2
            np.sum(fb[:, 19]),             # flt_bnk_eng20
            np.sum(fb[:, 12]),             # flt_bnk_eng13
            float(len(ss)),                # n_sil_segments
            sil_pct,                       # frac_silence
            min_sil,                       # min_sil_len
            jitter, shimmer, hnr,
        ]).reshape(1, -1)

        # [OPT] Clip to training range — vectorised (replaces Python for-loop)
        raw = np.clip(raw, self._clip_lo, self._clip_hi)

        print("done")
        return raw, ss, se, env

    # ── calibration ───────────────────────────────────────────────────────────
    def calibrate(self, features):
        cal = features.flatten().copy()
        for i, fn in enumerate(FEATURE_ORDER):
            mu, std, lo, hi = TRAINING_STATS[fn]
            z      = (cal[i] - mu) / (std + 1e-8)
            cal[i] = np.clip((z + INDIAN_ACCENT_OFFSETS.get(fn, 0.0)) * std + mu, lo, hi)
        return cal.reshape(1, -1)

    # ── prediction ────────────────────────────────────────────────────────────
    def predict(self, features):
        feats = self.calibrate(features) if self.indian_accent else features
        try:    fs = self.scaler.transform(feats)
        except: fs = feats
        proba = self.model.predict_proba(fs)[0]
        pred  = 1.0 if proba[1] >= self.confidence_threshold else 0.0
        return pred, proba, feats

    # ── ensemble (long audio) ─────────────────────────────────────────────────
    def _ensemble_predict(self, signal):
        seg_len = int(self.duration * self.sample_rate)
        starts  = np.linspace(0, max(0, len(signal) - seg_len), 2, dtype=int)
        preds, probas = [], []
        first_cal = None
        for i, s in enumerate(starts):
            f, *_ = self.extract_features(signal[s: s + seg_len])
            p, pr, cal = self.predict(f)
            preds.append(p); probas.append(pr)
            if i == 0: first_raw, first_cal = f, cal
        avg   = np.mean(probas, axis=0)
        final = 1.0 if sum(p == 1.0 for p in preds) > len(preds) / 2 else 0.0
        print(f"[ENSEMBLE] {sum(p==1 for p in preds)}/{len(preds)} segments → Dementia. Avg P(AD)={avg[1]*100:.1f}%")
        return final, avg, first_cal


    # ── waveform plot ─────────────────────────────────────────────────────────
    def plot_waveform(self, signal, ss, se, env, sil_pct=None):
        """Enhanced waveform plot with clear labels, legend, grid, and
        an automatic annotation when silence fraction is clinically high."""
        sr    = self.sample_rate
        t     = np.linspace(0, len(signal) / sr, len(signal))
        t_env = np.linspace(0, len(signal) / sr, len(env))

        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 6), facecolor='#0d1117')
        title_extra = ""
        if sil_pct is not None and sil_pct > 40:
            title_extra = f"  ⚠ High Silence: {sil_pct:.1f}% (normal < 40%)"
        fig.suptitle(f'Voice Waveform Analysis{title_extra}',
                     color='white', fontsize=13, fontweight='bold')

        for ax in (ax1, ax2):
            ax.set_facecolor('#161b22')
            for sp in ax.spines.values():
                sp.set_edgecolor('#30363d')
            ax.tick_params(colors='#8b949e')
            ax.grid(True, color='#30363d', alpha=0.5, linewidth=0.5)

        # ── Raw waveform
        ax1.plot(t, signal, color='#58a6ff', lw=0.5, alpha=0.85, label='Audio Signal')
        for s, e in zip(ss, se):
            ts = t_env[min(s, len(t_env)-1)]
            te = t_env[min(e, len(t_env)-1)]
            ax1.axvspan(ts, te, color='#f85149', alpha=0.20,
                        label='Silence Region' if s == ss[0] else '_')
        ax1.set_ylabel('Amplitude', color='#8b949e', fontsize=10)
        ax1.set_xlabel('Time (seconds)', color='#8b949e', fontsize=9)
        ax1.set_title('Raw Waveform  (red = silence regions)',
                      color='#c9d1d9', fontsize=10)
        ax1.legend(loc='upper right', facecolor='#1c2128',
                   labelcolor='#c9d1d9', fontsize=8)

        # ── Envelope
        ax2.plot(t_env, env, color='#3fb950', lw=0.8, label='Signal Envelope')
        ax2.fill_between(t_env, env, alpha=0.15, color='#3fb950')
        ax2.set_ylabel('Envelope Amplitude', color='#8b949e', fontsize=10)
        ax2.set_xlabel('Time (seconds)', color='#8b949e', fontsize=10)
        ax2.set_title('Signal Envelope — Voice Energy Over Time',
                      color='#c9d1d9', fontsize=10)
        ax2.legend(loc='upper right', facecolor='#1c2128',
                   labelcolor='#c9d1d9', fontsize=8)

        plt.tight_layout()
        out = os.path.join(self.output_dir, 'waveform_plot.png')
        fig.savefig(out, dpi=150, bbox_inches='tight', facecolor='#0d1117')
        plt.close(fig)
        print(f"[PLOT] Waveform → {os.path.abspath(out)}")

    # ── voice dashboard ───────────────────────────────────────────────────────
    def plot_voice_dashboard(self, signal, raw_f, cal_f, prediction, proba):
        """
        4-panel clinical voice analysis dashboard:
          1. Pitch (F0) variation over time — with abnormal zones highlighted
          2. Speech energy + vocal brightness (merged, shared time axis) —
             RMS energy with silent gaps shaded, spectral centroid overlaid
          3. Voice stability: jitter / shimmer / HNR vs AD/CN means
          4. Top-8 feature Z-score bar chart vs training distribution
        """
        sr = self.sample_rate
        BG, PANEL = '#0d1117', '#161b22'
        TITLE_C, TICK_C, SPINE_C = '#c9d1d9', '#8b949e', '#30363d'
        BLUE, GREEN, RED, GOLD = '#58a6ff', '#3fb950', '#f85149', '#ffd700'
        ORANGE = '#f0883e'

        fig = plt.figure(figsize=(18, 9), facecolor=BG)
        risk_str = ('ELEVATED RISK INDICATORS' if prediction == 1.0
                    else 'LOW RISK INDICATORS')
        risk_col  = RED if prediction == 1.0 else GREEN
        fig.suptitle(
            f'Voice Cognitive-Risk Dashboard  —  {risk_str}  '
            f'(P(AD)={proba[1]*100:.1f}%)',
            color=risk_col, fontsize=15, fontweight='bold', y=0.99)

        gs = plt.GridSpec(2, 2, figure=fig, hspace=0.45, wspace=0.30)

        def _style(ax, title, xlabel, ylabel):
            ax.set_facecolor(PANEL)
            for sp in ax.spines.values(): sp.set_edgecolor(SPINE_C)
            ax.tick_params(colors=TICK_C, labelsize=8)
            ax.set_title(title, color=TITLE_C, fontsize=10, fontweight='bold', pad=6)
            ax.set_xlabel(xlabel, color=TICK_C, fontsize=9)
            ax.set_ylabel(ylabel, color=TICK_C, fontsize=9)
            ax.grid(True, color=SPINE_C, alpha=0.6, linewidth=0.6)

        # ── 1. Pitch (F0) over time ──────────────────────────────────────────
        ax1 = fig.add_subplot(gs[0, 0])
        try:
            hop = 512
            f0 = librosa.yin(signal, fmin=librosa.note_to_hz('C2'),
                              fmax=librosa.note_to_hz('C7'),
                              sr=sr, frame_length=2048, hop_length=hop)
            times = librosa.frames_to_time(np.arange(len(f0)),
                                           sr=sr, hop_length=hop)
            voiced = f0 > 65
            f0_voiced = np.where(voiced, f0, np.nan)
            ax1.plot(times, f0_voiced, color=BLUE, lw=0.8, label='Pitch (F0)')
            # Highlight regions where pitch < 100 Hz (potential tremor / low energy)
            for i, (t, v, vv) in enumerate(zip(times, f0_voiced, voiced)):
                if vv and v < 100:
                    ax1.axvspan(t, min(t + hop/sr, times[-1]),
                                color=RED, alpha=0.15)
            f0_valid = f0_voiced[~np.isnan(f0_voiced)]
            if len(f0_valid):
                mean_f0 = float(np.mean(f0_valid))
                ax1.axhline(mean_f0, color=GOLD, lw=1.2, ls='--',
                            label=f'Mean F0 = {mean_f0:.0f} Hz')
            ax1.legend(facecolor='#1c2128', labelcolor=TITLE_C, fontsize=8,
                       loc='upper right')
        except Exception:
            ax1.text(0.5, 0.5, 'Pitch analysis unavailable',
                     ha='center', va='center', transform=ax1.transAxes,
                     color=TICK_C)
        _style(ax1, 'Pitch (F0) Variation Over Time',
               'Time (s)', 'Frequency (Hz)')

        # ── 2. Speech Energy + Vocal Brightness (merged, shared time axis) ────
        # RMS energy (when the person is speaking) and spectral centroid (how
        # bright/clear the voice is) share one time axis. Silent gaps -- where
        # there is no energy spike -- are shaded, so pauses and the brightness
        # during speech can be read together.
        ax2 = fig.add_subplot(gs[0, 1])
        try:
            hop_sr = 512
            rms_curve = librosa.feature.rms(y=signal, frame_length=2048,
                                             hop_length=hop_sr)[0]
            t_rms = librosa.frames_to_time(np.arange(len(rms_curve)),
                                           sr=sr, hop_length=hop_sr)
            sc = librosa.feature.spectral_centroid(y=signal, sr=sr,
                                                   hop_length=hop_sr)[0]
            t_sc = librosa.frames_to_time(np.arange(len(sc)),
                                          sr=sr, hop_length=hop_sr)

            sil_threshold = np.mean(rms_curve) * 0.5
            speaking = rms_curve > sil_threshold
            onsets = np.where(np.diff(speaking.astype(int)) == 1)[0]
            duration_s = len(signal) / sr
            speech_rate = len(onsets) / max(duration_s, 1)
            silence_frac = float(np.mean(~speaking))

            # Left axis: RMS energy (speech activity)
            ax2.plot(t_rms, rms_curve, color=GREEN, lw=0.9, label='Speech energy (RMS)')
            ax2.fill_between(t_rms, rms_curve, color=GREEN, alpha=0.15)
            ax2.axhline(sil_threshold, color=GREEN, lw=1.0, ls='--', alpha=0.7)
            ax2.set_ylim(0, max(rms_curve.max() * 1.1, 1e-3))

            # Shade the silent gaps (no energy spike)
            in_sil = False; start = 0
            for i, sp in enumerate(speaking):
                if not sp and not in_sil:
                    in_sil = True; start = t_rms[i]
                elif sp and in_sil:
                    in_sil = False
                    ax2.axvspan(start, t_rms[i], color=RED, alpha=0.10)
            if in_sil:
                ax2.axvspan(start, t_rms[-1], color=RED, alpha=0.10)

            # Right axis: spectral centroid (vocal brightness)
            ax2b = ax2.twinx()
            ax2b.plot(t_sc, sc, color=ORANGE, lw=0.8, alpha=0.85,
                      label='Vocal brightness (centroid)')
            mean_sc = float(np.mean(sc))
            ax2b.axhline(mean_sc, color=GOLD, lw=1.0, ls=':')
            ax2b.set_ylabel('Brightness (Hz)', color=ORANGE, fontsize=9)
            ax2b.tick_params(axis='y', colors=ORANGE, labelsize=8)
            for sp in ax2b.spines.values():
                sp.set_edgecolor(SPINE_C)

            rate_label = ('Slow' if speech_rate < 2 else
                          'Normal' if speech_rate <= 4 else 'Fast')
            # combine legends from both axes
            h1, l1 = ax2.get_legend_handles_labels()
            h2, l2 = ax2b.get_legend_handles_labels()
            ax2.legend(h1 + h2, l1 + l2, facecolor='#1c2128',
                       labelcolor=TITLE_C, fontsize=7.5, loc='upper right')
            title2 = (f'Speech Energy & Vocal Brightness  —  '
                      f'{speech_rate:.1f} syl/s ({rate_label}), '
                      f'{silence_frac*100:.0f}% silence')
        except Exception:
            title2 = 'Speech Energy & Vocal Brightness'
            ax2.text(0.5, 0.5, 'Speech/brightness analysis unavailable',
                     ha='center', va='center', transform=ax2.transAxes,
                     color=TICK_C)
        _style(ax2, title2, 'Time (s)', 'RMS Energy')

        # ── 3. Voice Stability — jitter / shimmer / HNR bar chart ────────────
        ax5 = fig.add_subplot(gs[1, 0])
        cal = cal_f.flatten()
        fi  = self._feat_idx
        metrics = {
            'Jitter':  (cal[fi['jitter']],
                        AD_MEANS['jitter'], CN_MEANS['jitter'],
                        TRAINING_STATS['jitter'][0]),
            'Shimmer': (cal[fi['shimmer']],
                        AD_MEANS['shimmer'], CN_MEANS['shimmer'],
                        TRAINING_STATS['shimmer'][0]),
            'HNR':     (cal[fi['HNR']],
                        AD_MEANS['HNR'], CN_MEANS['HNR'],
                        TRAINING_STATS['HNR'][0]),
        }
        labels_v = list(metrics.keys())
        patient_v = [metrics[k][0] for k in labels_v]
        ad_v      = [metrics[k][1] for k in labels_v]
        cn_v      = [metrics[k][2] for k in labels_v]
        x = np.arange(len(labels_v)); w = 0.26
        ax5.bar(x - w, patient_v, w, color=ORANGE,  label='Patient',  alpha=0.9)
        ax5.bar(x,     ad_v,      w, color=RED,     label='AD avg',   alpha=0.75)
        ax5.bar(x + w, cn_v,      w, color=GREEN,   label='CN avg',   alpha=0.75)
        ax5.set_xticks(x)
        ax5.set_xticklabels(labels_v, color=TICK_C, fontsize=9)
        ax5.legend(facecolor='#1c2128', labelcolor=TITLE_C, fontsize=8)
        _style(ax5, 'Voice Stability Pattern (vs AD/CN Reference)',
               'Metric', 'Value')

        # ── 4. Top-8 Feature Z-scores ─────────────────────────────────────────
        ax6 = fig.add_subplot(gs[1, 1])
        raw = raw_f.flatten()
        z_scores = np.array([
            (cal[fi[fn]] - TRAINING_STATS[fn][0]) / (TRAINING_STATS[fn][1] + 1e-8)
            for fn in EXCEL_FEATURES
        ])
        colors_z = [RED if abs(z) > 1.5 else (GOLD if abs(z) > 0.8 else GREEN)
                    for z in z_scores]
        x_z = np.arange(len(EXCEL_FEATURES))
        bars = ax6.bar(x_z, z_scores, color=colors_z, alpha=0.85, width=0.65)
        ax6.axhline(0,    color=SPINE_C, lw=0.8, ls='--')
        ax6.axhline( 1.5, color=RED,    lw=0.8, ls=':', label='±1.5σ (abnormal)')
        ax6.axhline(-1.5, color=RED,    lw=0.8, ls=':')
        ax6.set_xticks(x_z)
        ax6.set_xticklabels(EXCEL_FEATURES, rotation=30, ha='right',
                            color=TICK_C, fontsize=7.5)
        ax6.legend(facecolor='#1c2128', labelcolor=TITLE_C, fontsize=8)
        _style(ax6, 'Top-8 Discriminant Features — Z-score vs Training',
               'Feature', 'Z-score (σ)')

        out = os.path.join(self.output_dir, 'voice_dashboard.png')
        fig.savefig(out, dpi=150, bbox_inches='tight', facecolor=BG)
        plt.close(fig)
        print(f"[PLOT] Voice dashboard → {os.path.abspath(out)}")
        return os.path.abspath(out)

    # ── feature plot ──────────────────────────────────────────────────────────
    def plot_features(self, raw_f, cal_f, prediction, proba):
        raw = raw_f.flatten(); cal = cal_f.flatten()
        idx = self._feat_idx

        fig = plt.figure(figsize=(18, 13), facecolor='#0d1117')
        fig.suptitle('Feature Analysis — Voice Dementia Screening',
                     color='white', fontsize=14, fontweight='bold', y=0.98)
        gs = gridspec.GridSpec(2, 2, figure=fig, hspace=0.50, wspace=0.35)

        ax1 = fig.add_subplot(gs[0, :])
        ax1.set_facecolor('#161b22')
        for sp in ax1.spines.values(): sp.set_edgecolor('#30363d')
        ax1.tick_params(colors='#8b949e')
        ax1.grid(True, color='#30363d', alpha=0.5, linewidth=0.5)

        x = np.arange(len(FEATURE_ORDER))
        def zs(vals):
            return np.array([(vals[i] - TRAINING_STATS[fn][0]) / (TRAINING_STATS[fn][1] + 1e-8)
                             for i, fn in enumerate(FEATURE_ORDER)])
        raw_z = zs(raw); cal_z = zs(cal)

        w = 0.38
        ax1.bar(x - w/2, raw_z, w, label='Raw',        color='#58a6ff', alpha=0.75)
        ax1.bar(x + w/2, cal_z, w, label='Calibrated', color='#3fb950', alpha=0.75)
        ax1.axhline(0,  color='#8b949e', lw=0.8, ls='--')
        ax1.axhline( 2, color='#f85149', lw=0.6, ls=':', label='±2σ (abnormal)')
        ax1.axhline(-2, color='#f85149', lw=0.6, ls=':')
        ax1.set_xticks(x)
        ax1.set_xticklabels(FEATURE_ORDER, rotation=45, ha='right',
                            fontsize=7.5, color='#8b949e')
        ax1.set_ylabel('Z-score (σ)', color='#8b949e', fontsize=10)
        ax1.set_xlabel('Acoustic Feature', color='#8b949e', fontsize=9)
        ax1.set_title('All 21 Features — Z-score vs Training Distribution  '
                      '(gold highlight = saved to Excel)',
                      color='#c9d1d9', fontsize=10)
        ax1.legend(facecolor='#1c2128', labelcolor='#c9d1d9', fontsize=8)
        for fn in EXCEL_FEATURES:
            ax1.axvspan(idx[fn] - 0.5, idx[fn] + 0.5, color='#ffd700', alpha=0.07)

        ax2 = fig.add_subplot(gs[1, 0], polar=True)
        ax2.set_facecolor('#1c2128')
        labels = EXCEL_FEATURES
        n      = len(labels)
        angles = np.linspace(0, 2 * np.pi, n, endpoint=False).tolist() + [0]

        def norm_vals(feat_list, source):
            out = []
            for fn in feat_list:
                i = idx[fn]; mu, std, lo, hi = TRAINING_STATS[fn]
                z    = (source[i] - mu) / (std + 1e-8)
                z_lo = (lo - mu) / (std + 1e-8); z_hi = (hi - mu) / (std + 1e-8)
                out.append(float(np.clip((z - z_lo) / (z_hi - z_lo + 1e-8), 0, 1)))
            return out + [out[0]]

        ad_src = [AD_MEANS[fn] for fn in FEATURE_ORDER]
        cn_src = [CN_MEANS[fn] for fn in FEATURE_ORDER]

        ax2.plot(angles, norm_vals(labels, cal),    'o-', lw=2, color='#f0883e', label='Patient')
        ax2.fill(angles, norm_vals(labels, cal),    alpha=0.15, color='#f0883e')
        ax2.plot(angles, norm_vals(labels, ad_src), 's--', lw=1.2, color='#f85149', label='AD avg')
        ax2.plot(angles, norm_vals(labels, cn_src), '^--', lw=1.2, color='#3fb950', label='CN avg')
        ax2.set_xticks(angles[:-1])
        ax2.set_xticklabels(labels, fontsize=7.5, color='#c9d1d9')
        ax2.set_title('Top-8 Discriminant Features\n(normalised radar)',
                      color='#c9d1d9', fontsize=10, pad=15)
        ax2.legend(loc='upper right', bbox_to_anchor=(1.35, 1.1),
                   facecolor='#1c2128', labelcolor='#c9d1d9', fontsize=8)

        ax3 = fig.add_subplot(gs[1, 1])
        ax3.set_facecolor('#161b22'); ax3.axis('off')
        for sp in ax3.spines.values(): sp.set_edgecolor('#30363d')

        rc = '#f85149' if prediction == 1.0 else '#3fb950'
        rt = 'DEMENTIA DETECTED' if prediction == 1.0 else 'NO DEMENTIA DETECTED'
        ax3.text(0.5, 0.90, rt, transform=ax3.transAxes,
                 ha='center', fontsize=15, fontweight='bold', color=rc)
        ax3.text(0.5, 0.76, f'P(Dementia) = {proba[1]*100:.1f}%',
                 transform=ax3.transAxes, ha='center', fontsize=12, color='#f85149')
        ax3.text(0.5, 0.65, f'P(Normal)   = {proba[0]*100:.1f}%',
                 transform=ax3.transAxes, ha='center', fontsize=12, color='#3fb950')
        ax3.text(0.5, 0.54, f'Threshold: ≥{self.confidence_threshold*100:.0f}% → Dementia',
                 transform=ax3.transAxes, ha='center', fontsize=9, color='#8b949e')

        ax3.text(0.5, 0.42, '── Top Feature Deviations ──',
                 transform=ax3.transAxes, ha='center', fontsize=8, color='#8b949e')
        for ri, fn in enumerate(EXCEL_FEATURES[:6]):
            fi = idx[fn]; mu, std, *_ = TRAINING_STATS[fn]
            z  = (cal[fi] - mu) / (std + 1e-8)
            c  = '#f85149' if z > 1.5 else ('#ffd700' if z > 0.8 else '#3fb950')
            ax3.text(0.5, 0.34 - ri * 0.048,
                     f'{fn:<16s}  z={z:+.2f}',
                     transform=ax3.transAxes, ha='center',
                     fontsize=8, color=c, family='monospace')

        ax3.text(0.5, 0.01,
                 '⚠  Screening tool only — not a medical diagnosis.\n'
                 'Consult a neurologist for clinical evaluation.',
                 transform=ax3.transAxes, ha='center', fontsize=7.5,
                 color='#8b949e', style='italic')

        fig.savefig(os.path.join(self.output_dir, 'feature_plot.png'),
                    dpi=150, bbox_inches='tight', facecolor='#0d1117')
        plt.close(fig)
        print(f"[PLOT] Feature chart → {os.path.abspath(os.path.join(self.output_dir, 'feature_plot.png'))}")

    # ── excel ─────────────────────────────────────────────────────────────────
    def save_to_excel(self, cal_f, prediction, proba, source, path='dementia_features_log.xlsx'):
        if not OPENPYXL_AVAILABLE:
            print("[SKIP] pip install openpyxl"); return
        cal = cal_f.flatten(); idx = self._feat_idx

        header = (['Timestamp', 'Audio_Source', 'Prediction', 'Confidence_%',
                   'P_Dementia_%', 'P_Normal_%'] + EXCEL_FEATURES + ['Notes'])
        pred_label = 'Dementia' if prediction == 1.0 else 'No Dementia'
        note = ('Borderline — repeat recommended' if prediction == 1.0 and proba[1] < 0.75
                else 'High confidence — clinical eval advised' if prediction == 1.0 else '')
        row = ([datetime.now().strftime('%Y-%m-%d %H:%M:%S'), source, pred_label,
                round(float(max(proba)) * 100, 2),
                round(float(proba[1]) * 100, 2),
                round(float(proba[0]) * 100, 2)]
               + [round(float(cal[idx[fn]]), 6) for fn in EXCEL_FEATURES]
               + [note])

        try:
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path); ws = wb.active
            else:
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Dementia Screening'
                ws.append(header)
                ws.append(['', "Cohen's d effect sizes (higher = better discriminant)", '', '', '', '']
                          + [FEATURE_DESCRIPTIONS.get(fn, fn) for fn in EXCEL_FEATURES] + [''])
                hf = PatternFill('solid', fgColor='1F4E79')
                ff = PatternFill('solid', fgColor='2E4057')
                for ci, c in enumerate(ws[1], 1):
                    c.fill = hf if ci <= 6 else ff
                    c.font = Font(bold=True, color='FFFFFF')
                    c.alignment = Alignment(horizontal='center')
            ws.append(row)
            from openpyxl.styles import Font, PatternFill
            pc = ws.cell(ws.max_row, 3)
            if prediction == 1.0:
                pc.fill = PatternFill('solid', fgColor='F4CCCC')
                pc.font = Font(bold=True, color='C00000')
            else:
                pc.fill = PatternFill('solid', fgColor='D9EAD3')
                pc.font = Font(bold=True, color='274E13')
            wb.save(path)
            print(f"[EXCEL] Saved → {os.path.abspath(path)}")
        except Exception as e:
            print(f"[WARN] Excel save failed: {e}")

    # ── main run ──────────────────────────────────────────────────────────────
    def run(self, audio_file=None, plot=True, show_features=False,
            excel_path=None, ensemble=True):
        print("\n" + "=" * 65)
        print("  DEMENTIA ANALYSIS SYSTEM  v4.1  (Fast Edition)")
        if self.indian_accent:
            print("  Mode: Indian Accent Optimised")
        print(f"  Confidence threshold: {self.confidence_threshold*100:.0f}%")
        print("=" * 65)

        if audio_file:
            if not os.path.exists(audio_file):
                raise FileNotFoundError(audio_file)
            signal, _ = librosa.load(audio_file, sr=self.sample_rate)
            source = os.path.basename(audio_file)
            print(f"[FILE] {audio_file}  ({len(signal)/self.sample_rate:.1f}s)")
        else:
            signal = self.record_audio()
            source = 'microphone'

        if ensemble and len(signal) > 30 * self.sample_rate:
            print("[INFO] Long audio — running 2-segment ensemble")
            prediction, proba, cal_f = self._ensemble_predict(signal)
            raw_f, ss, se, env = self.extract_features(signal[:int(self.duration * self.sample_rate)])
        else:
            raw_f, ss, se, env = self.extract_features(signal)
            prediction, proba, cal_f = self.predict(raw_f)

        # Compute silence % for waveform annotation
        env_temp = self._envelope(signal)
        ss_t, se_t, _ = self._silence_detection(env_temp)
        sil_pct_val = float(np.sum(se_t - ss_t)) / max(len(env_temp), 1) * 100

        print("\n" + "=" * 65)
        if prediction == 1.0:
            print(f"\n  [!] DEMENTIA SIGNS DETECTED   P={proba[1]*100:.1f}%")
        else:
            print(f"\n  [OK] NO DEMENTIA SIGNS        P(Normal)={proba[0]*100:.1f}%")
        print(f"\n  P(Dementia)={proba[1]*100:.1f}%  P(Normal)={proba[0]*100:.1f}%")
        print(f"  Threshold  = {self.confidence_threshold*100:.0f}%  (↑ from 50% to cut false positives)")

        if show_features:
            print(f"\n  {'':2} {'Feature':<18} {'Raw':>12} {'Calibrated':>12} {'AD-avg':>10} {'CN-avg':>10}")
            print("  " + "-" * 62)
            for i, fn in enumerate(FEATURE_ORDER):
                mk = '**' if fn in EXCEL_FEATURES else '  '
                print(f"  {mk}{fn:<18} {raw_f.flatten()[i]:>12.4f}"
                      f" {cal_f.flatten()[i]:>12.4f}"
                      f" {AD_MEANS[fn]:>10.4f} {CN_MEANS[fn]:>10.4f}")
            print("  ** = saved to Excel")

        print("\n  DISCLAIMER: Screening tool — not a medical diagnosis.")
        print("=" * 65 + "\n")

        if plot:
            self.plot_waveform(signal, ss, se, env, sil_pct=sil_pct_val)
            self.plot_features(raw_f, cal_f, prediction, proba)
            self.plot_voice_dashboard(signal, raw_f, cal_f, prediction, proba)

        ep = excel_path or os.path.join(self.output_dir, 'dementia_features_log.xlsx')
        self.save_to_excel(cal_f, prediction, proba, source, ep)

        if plot:
            try:
                plt.show(block=True)
            except Exception:
                pass

        return prediction, proba, raw_f, cal_f

    # ── GUI-callable API ───────────────────────────────────────────────────────
    def run_analysis_for_gui(self, audio_file=None, plot=True,
                              excel_path=None, ensemble=True,
                              progress_callback=None):
        """
        Run the full dementia analysis pipeline and return a structured result
        dictionary suitable for display in a GUI.

        Parameters
        ----------
        audio_file      : str | None   — path to WAV file, or None for mic
        plot            : bool          — whether to save plot images
        excel_path      : str | None   — override Excel output path
        ensemble        : bool          — use 2-segment ensemble for long audio
        progress_callback : callable | None
                          — called with (step: str, pct: int) during processing

        Returns
        -------
        dict with keys:
            prediction      int    0 = no dementia, 1 = dementia detected
            proba           list   [P(normal), P(dementia)]  0–1 floats
            risk_pct        float  P(dementia) × 100
            risk_category   str    'Low' | 'Moderate' | 'High'
            key_indicators  list   human-readable warning strings
            metrics         dict   jitter, shimmer, HNR, frac_silence,
                                   speech_rate, pitch_variability
            plot_paths      dict   waveform, feature, dashboard PNG paths
            excel_path      str    path to Excel log (or None)
            error           str    non-empty if something went wrong
        """

        def _cb(step, pct):
            if progress_callback:
                try:
                    progress_callback(step, pct)
                except Exception:
                    pass

        result = {
            'prediction': 0, 'proba': [1.0, 0.0], 'risk_pct': 0.0,
            'risk_category': 'Low', 'key_indicators': [],
            'metrics': {}, 'plot_paths': {}, 'excel_path': None,
            'error': ''
        }

        try:
            _cb('Loading audio…', 5)
            if audio_file:
                if not os.path.exists(audio_file):
                    result['error'] = f"Audio file not found: {audio_file}"
                    return result
                signal, _ = librosa.load(audio_file, sr=self.sample_rate)
                source = os.path.basename(audio_file)
            else:
                _cb('Recording from microphone…', 10)
                signal = self.record_audio()
                source = 'microphone'

            _cb('Extracting acoustic features…', 30)
            if ensemble and len(signal) > 30 * self.sample_rate:
                prediction, proba, cal_f = self._ensemble_predict(signal)
                raw_f, ss, se, env = self.extract_features(
                    signal[:int(self.duration * self.sample_rate)])
            else:
                raw_f, ss, se, env = self.extract_features(signal)
                prediction, proba, cal_f = self.predict(raw_f)

            _cb('Computing voice metrics…', 55)
            cal = cal_f.flatten()
            fi  = self._feat_idx

            # Compute silence %
            sil_pct_val = float(np.sum(se - ss)) / max(len(env), 1) * 100

            # Estimate speech rate
            try:
                hop_sr = 512
                rms_c = librosa.feature.rms(
                    y=signal, frame_length=2048, hop_length=hop_sr)[0]
                thr   = np.mean(rms_c) * 0.5
                above = rms_c > thr
                onsets = np.where(np.diff(above.astype(int)) == 1)[0]
                speech_rate = len(onsets) / max(len(signal) / self.sample_rate, 1)
            except Exception:
                speech_rate = 0.0

            # Pitch variability (std of voiced F0)
            try:
                f0 = librosa.yin(
                    signal[:int(15 * self.sample_rate)],
                    fmin=librosa.note_to_hz('C2'),
                    fmax=librosa.note_to_hz('C7'),
                    sr=self.sample_rate,
                    frame_length=2048, hop_length=512)
                pitch_var = float(np.std(f0[f0 > 65])) if np.any(f0 > 65) else 0.0
            except Exception:
                pitch_var = 0.0

            metrics = {
                'jitter':          round(float(cal[fi['jitter']]),   5),
                'shimmer':         round(float(cal[fi['shimmer']]),  4),
                'HNR':             round(float(cal[fi['HNR']]),      2),
                'frac_silence':    round(sil_pct_val,                1),
                'speech_rate':     round(speech_rate,                2),
                'pitch_variability': round(pitch_var,                2),
            }

            # Key indicators
            indicators = []
            if speech_rate < 2.0:
                indicators.append("Slower-than-normal speech rate")
            if pitch_var < 15.0:
                indicators.append("Reduced pitch variability (monotone speech)")
            if float(cal[fi['jitter']]) > TRAINING_STATS['jitter'][0] + TRAINING_STATS['jitter'][1]:
                indicators.append("Elevated voice tremor (jitter)")
            if float(cal[fi['shimmer']]) > TRAINING_STATS['shimmer'][0] + TRAINING_STATS['shimmer'][1]:
                indicators.append("High amplitude instability (shimmer)")
            if float(cal[fi['HNR']]) < TRAINING_STATS['HNR'][0] - TRAINING_STATS['HNR'][1]:
                indicators.append("Poor harmonics-to-noise ratio (breathy voice)")
            if sil_pct_val > 40:
                indicators.append(f"Excessive silence fraction ({sil_pct_val:.0f}% > 40%)")

            risk_pct = proba[1] * 100
            if risk_pct < 40:
                risk_cat = 'Low'
            elif risk_pct < 65:
                risk_cat = 'Moderate'
            else:
                risk_cat = 'High'

            # Approximate per-feature attribution for explainability (guarded)
            feature_contributions = _compute_voice_attribution(self.model, cal)

            _cb('Generating plots…', 70)
            plot_paths = {}
            if plot:
                self.plot_waveform(signal, ss, se, env, sil_pct=sil_pct_val)
                plot_paths['waveform'] = os.path.abspath(
                    os.path.join(self.output_dir, 'waveform_plot.png'))
                self.plot_features(raw_f, cal_f, float(prediction), proba)
                plot_paths['feature'] = os.path.abspath(
                    os.path.join(self.output_dir, 'feature_plot.png'))
                dashboard = self.plot_voice_dashboard(
                    signal, raw_f, cal_f, float(prediction), proba)
                plot_paths['dashboard'] = dashboard

            _cb('Saving Excel log…', 88)
            ep = excel_path or os.path.join(self.output_dir, 'dementia_features_log.xlsx')
            self.save_to_excel(cal_f, float(prediction), proba, source, ep)

            # Capture full feature vector for the (label-driven) retraining loop
            log_session_for_retraining(
                cal_f.flatten(), float(prediction), proba, source,
                os.path.join(self.output_dir, 'retrain_data.csv'))

            result.update({
                'prediction':     int(prediction),
                'proba':          [round(float(p), 4) for p in proba],
                'risk_pct':       round(risk_pct, 1),
                'risk_category':  risk_cat,
                'key_indicators': indicators,
                'metrics':        metrics,
                'plot_paths':     plot_paths,
                'excel_path':     ep,
                'feature_contributions': feature_contributions,
            })
            _cb('Done', 100)

        except Exception as exc:
            import traceback
            result['error'] = str(exc)
            traceback.print_exc()

        return result


# ─────────────────────────────────────────────────────────────────────────────
def main():
    import argparse
    p = argparse.ArgumentParser(description='Dementia Analysis v4.1 — Fast Edition')
    p.add_argument('--audio',            type=str)
    p.add_argument('--model',            type=str,  default='dementia_rf_model.pkl')
    p.add_argument('--scaler',           type=str,  default='scaler.pkl')
    p.add_argument('--plot',             action='store_true')
    p.add_argument('--show-features',    action='store_true')
    p.add_argument('--no-indian-accent', action='store_true')
    p.add_argument('--no-ensemble',      action='store_true')
    p.add_argument('--threshold',        type=float, default=0.60)
    p.add_argument('--excel',            type=str,  default=None)
    p.add_argument('--output-dir',       type=str,  default='.')
    args = p.parse_args()

    print(f"[MODE] {'File: ' + args.audio if args.audio else 'Microphone'} | "
          f"Indian accent: {'ON' if not args.no_indian_accent else 'OFF'} | "
          f"Threshold: {args.threshold*100:.0f}%")
    try:
        a = DementiaAnalyser(model_path=args.model, scaler_path=args.scaler,
                             indian_accent=not args.no_indian_accent,
                             confidence_threshold=args.threshold,
                             output_dir=args.output_dir)
        a.run(audio_file=args.audio, plot=args.plot,
              show_features=args.show_features,
              ensemble=not args.no_ensemble, excel_path=args.excel)
    except KeyboardInterrupt:
        print("\nCancelled.")
    except Exception as e:
        import traceback; print(f"\n[ERROR] {e}"); traceback.print_exc()


if __name__ == '__main__':
    main()

